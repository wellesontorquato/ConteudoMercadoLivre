from flask import Flask, render_template, send_file, make_response
from flask import request as req
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
import traceback
import re

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download_planilha', methods=['POST'])
def download_planilha():
    try:
        file = req.files['html_file']

        # Ler o conteúdo do arquivo HTML
        html_content = file.read().decode('utf-8')

        # Analisar o conteúdo do HTML com BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')
        numero_envio = soup.select_one('.inbound__flow-title-cx span').text
        numero_envio = re.search(r'\#(\d+)', numero_envio).group(1)
        lista_produtos = soup.find_all('tr', {'class': 'andes-table__row product-row false'})

        # Criar a planilha e definir suas propriedades
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Produtos"

        sheet.column_dimensions['A'].width = 61
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15

        default_font = Font(name='Arial', size=10)

        # Configurar as propriedades do cabeçalho
        header_labels = ["DESCRIÇÃO", "MLB", "SKU", "CRIADO", "ENVIADO", "DATA", "ENVIO"]
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="000000")
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Cabeçalho da planilha
        for label in header_labels:
            cell = sheet.cell(row=1, column=header_labels.index(label) + 1)
            cell.value = label
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = header_border

        # Extrair os dados do HTML
        for produto in lista_produtos:
            descricao = produto.find('h5', {'class': 'unit-row__title'}).text
            sku = produto.find('dt', {'class': 'unit-row__details-value unit-row__details-value--code'}).find(
                'span').text
            mlb = produto.find('dt', {'class': 'unit-row__details-label'}).text
            mlb = int(re.findall(r'\d+', mlb)[0])
            mlbconcatenado = ''.join(['MLB', str(mlb)])
            quantidade = int(produto.find('input', {'type': 'number'})['value'])
            sheet.append([descricao, mlbconcatenado, sku, quantidade, "", "", numero_envio])  # Deixe a coluna vazia para a data

            # Aplicar formatação às células com dados
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Inserir o número de envio na célula correspondente
        numero_envio_cell = sheet.cell(row=2, column=7)  # Supondo que o número de envio seja o primeiro na lista de produtos
        numero_envio_cell.value = numero_envio

        # Mesclar células na coluna E (coluna 5) com base na quantidade de células preenchidas na coluna D (coluna 4)
        qtd_celulas_preenchidas_d = sheet.max_row - 1  # Desconsiderando a linha do cabeçalho
        if qtd_celulas_preenchidas_d > 1:
            start_row = 2  # Linha inicial da mesclagem
            end_row = start_row + qtd_celulas_preenchidas_d - 1  # Linha final da mesclagem

            # Mesclar as células
            sheet.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
            sheet.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

            # Preencher a célula mesclada com a data atual
            data_atual = date.today().strftime("%d/%m")
            cell_mesclada = sheet.cell(row=start_row, column=6)
            cell_mesclada.value = data_atual

        # Centralizando o conteúdo das colunas C e D
        for column in ['B', 'C', 'D', 'E']:
            for cell in sheet[column]:
                cell.alignment = Alignment(horizontal='center')

        # Centralizando o conteúdo das colunas E e F horizontal e vertical
        for column in ['F', 'G']:
            for cell in sheet[column]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Remover as linhas da planilha
        sheet.sheet_view.showGridLines = False

        # Definir o conteúdo da planilha como Arial 10px
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = default_font

        # Salvar a planilha
        workbook.save("produtos.xlsx")

        # Criar uma resposta para o arquivo
        response = make_response(send_file("produtos.xlsx", as_attachment=True))
        response.headers["Content-Disposition"] = "attachment; filename=produtos_a_planilhar.xlsx"

        return response

    except Exception as e:
        traceback.print_exc()
        return {"error": "Ocorreu um erro durante o processamento da solicitação."}


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)
